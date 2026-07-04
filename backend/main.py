import json
import os
import sqlite3
import sys

# Add the project root to sys.path so 'backend.*' imports work when run as a script
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm

import backend.utils.db_writer as db_writer
from backend.agents.research_agent import ResearchAgent, ensure_sources_field
from backend.config import settings
from backend.utils.db_writer import (
    create_run,
    fetch_all,
    get_all_existing_ids,
    get_db_path,
    init_db,
    prepare_row_data,
    save_results_bulk,
)
from backend.utils.jobs import (
    get_job,
    update_job_progress,
    update_job_status,
    update_job_token_usage,
    update_job_total_items,
)
from backend.utils.schemas import TokenUsage

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.FileHandler("collector.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


def format_output_excel(filepath: str, df: pd.DataFrame | None) -> None:
    if df is None or df.empty:
        logger.warning("No data to save to Excel.")
        return

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        column_idx = col[0].column
        if not isinstance(column_idx, int):
            continue

        max_length = 0
        col_letter = get_column_letter(column_idx)
        for cell in col:
            cell.alignment = wrap_alignment
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length * 1.1, 60)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(filepath)
    logger.info(f"Results formatted and saved to {filepath}")


def main(
    job_id: str | None = None,
    input_file: str | None = None,
    output_file: str | None = None,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> None:
    logger.info("Starting AI Data Collector")

    input_file_path = input_file or settings.input_file
    output_file_path = output_file or settings.output_file

    if job_id:
        job = get_job(job_id)
        if job and job.get("status") == "cancelled":
            return
        update_job_status(job_id, "running")

    job_record = get_job(job_id) if job_id else None

    effective_fields_raw = job_record.get("target_fields") if job_record else None
    try:
        effective_fields = (
            json.loads(effective_fields_raw) if effective_fields_raw else settings.target_fields
        )
    except Exception:
        effective_fields = settings.target_fields

    effective_item_label = (
        job_record.get("item_label") if job_record else None
    ) or settings.item_label

    output_fields = ensure_sources_field(effective_fields)

    effective_column = column_name or settings.column_name

    current_run_id = None
    if job_id:
        # Initialize schema without creating a default run
        init_db(output_fields, create_default_run=False, identifier_column=effective_column)
        # Create a specific run for this job
        current_run_id = create_run(
            input_file_path,
            output_file_path,
            settings.model_name,
            settings.web_search_provider,
        )
        # Link run_id to job
        conn = sqlite3.connect(get_db_path())
        try:
            conn.execute("UPDATE jobs SET run_id = ? WHERE job_id = ?", (current_run_id, job_id))
            conn.commit()
        except Exception as e:
            logger.error(f"Failed to associate run_id with job: {e}")
        finally:
            conn.close()
    else:
        # Normal CLI flow
        init_db(output_fields, identifier_column=effective_column)
        current_run_id = db_writer._CURRENT_RUN_ID

    effective_sheet = sheet_name or settings.sheet_name

    try:
        if input_file_path.endswith(".csv"):
            df = pd.read_csv(input_file_path)
        else:
            df = pd.read_excel(input_file_path, sheet_name=effective_sheet)
        if job_id:
            update_job_total_items(job_id, len(df))
    except Exception as e:
        logger.error(f"Failed to read input file: {e}")
        if job_id:
            update_job_status(job_id, "failed", str(e))
        return

    agent = ResearchAgent(item_label=effective_item_label)
    buffer: list[tuple[str, ...]] = []
    batch_confidence: list[dict[str, float | None]] = []
    batch_token_usage: list[TokenUsage] = []
    existing_ids = get_all_existing_ids(identifier_column=effective_column)
    try:
        col_idx = list(df.columns).index(effective_column) + 1
    except ValueError:
        error_msg = f"Column '{effective_column}' not found in the input file."
        logger.error(error_msg)
        if job_id:
            update_job_status(job_id, "failed", error_msg)
        return

    try:
        for start in range(0, len(df), settings.batch_size):
            if job_id:
                current_job = get_job(job_id)
                if current_job and current_job.get("status") == "cancelled":
                    logger.info(f"Job {job_id} was cancelled — stopping after batch {start}")
                    return

            batch_df = df.iloc[start : start + settings.batch_size]

            processed_in_batch = 0
            skipped_in_batch = 0
            failed_in_batch = 0
            batch_usage = TokenUsage()

            for row in tqdm(batch_df.itertuples(), total=len(batch_df), desc="Processing batch"):
                item_id = str(row[col_idx])

                if item_id in existing_ids:
                    logger.debug(f"Skipping {item_id} — already in database")
                    skipped_in_batch += 1
                    continue

                try:
                    parsed, conf, item_usage = agent.collect_item_with_confidence(
                        item_id,
                        output_fields,
                    )
                    row_data = prepare_row_data(
                        item_id, parsed, output_fields, identifier_column=effective_column
                    )
                    buffer.append(row_data)
                    batch_confidence.append(conf)
                    batch_token_usage.append(item_usage)
                    batch_usage = batch_usage + item_usage
                    processed_in_batch += 1
                except Exception as e:
                    logger.error(f"Failed processing item {item_id}: {e}")
                    failed_in_batch += 1

            if buffer:
                save_results_bulk(
                    buffer,
                    output_fields,
                    run_id=current_run_id,
                    confidence_list=batch_confidence,
                    token_usage_list=batch_token_usage,
                    identifier_column=effective_column,
                )
                buffer.clear()
                batch_confidence.clear()
                batch_token_usage.clear()

            if job_id:
                update_job_progress(
                    job_id,
                    processed=processed_in_batch,
                    skipped=skipped_in_batch,
                    failed=failed_in_batch,
                )
                if batch_usage.effective_llm_requests:
                    update_job_token_usage(
                        job_id,
                        batch_usage.prompt_tokens,
                        batch_usage.completion_tokens,
                        batch_usage.estimated_cost_usd,
                        batch_usage.effective_llm_requests,
                    )

        final_df = fetch_all(
            run_id=current_run_id if job_id else None, identifier_column=effective_column
        )
        format_output_excel(output_file_path, final_df)
        logger.info("Data collection completed successfully")
        if job_id:
            update_job_status(job_id, "completed")

    except Exception as e:
        logger.error(f"Data collection failed: {e}")
        if job_id:
            update_job_status(job_id, "failed", str(e))


if __name__ == "__main__":
    main()
